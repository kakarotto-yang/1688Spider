import os
import random
import time
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
import pandas as pd
from selenium.webdriver.support.wait import WebDriverWait

from setting import cur_dir, PAGE
from util.browserUtil import create_chrome_driver, add_cookies
from selenium.webdriver.support import expected_conditions as EC

driver = create_chrome_driver()
IMG_COL = 'A'


def createDir():
    if os.path.exists(cur_dir + '\data'):
        pass
    else:
        os.mkdir(cur_dir + '\data')
    if os.path.exists(cur_dir + '\done'):
        pass
    else:
        os.mkdir(cur_dir + '\done')
    if os.path.exists(cur_dir + '\img'):
        pass
    else:
        os.mkdir(cur_dir + '\img')
    if os.path.exists(cur_dir + '\imgtmp'):
        pass
    else:
        os.mkdir(cur_dir + '\imgtmp')


def createXls(path):
    if os.path.exists(path):
        return
    wb = Workbook()
    cols = ['首图', '价格', '销量', '链接', '公司名称', '经营模式']
    sheet = wb['Sheet']
    for col in cols:
        sheet.cell(1, cols.index(col) + 1, col)
    sheet.column_dimensions['A'].width = 14
    wb.save(path)


# 从网络下载首图
def save_img(url):
    res = requests.get(url)
    file_name = url.split('/')[-1]
    file_name = file_name[:file_name.index('?')]
    imgpath = cur_dir + '\imgtmp\\' + file_name
    with open(imgpath, 'wb') as f:
        for data in res.iter_content(128):
            f.write(data)
    return imgpath


# 插入图片到excel
def insert_img(sh, row, imgpath):
    sh.row_dimensions[row].height = 80
    img = Image(imgpath)
    img.width, img.height = 100, 100
    sh.add_image(img, 'A' + str(row))


# 保存首图等
def saveTmpData(path, data):
    try:
        wb = load_workbook(path)
        sheet = wb['Sheet']
        rows = sheet.max_row + 1
        isInsertImg = False
        for d in data:
            if not isInsertImg:
                imgpath = save_img(d)
                insert_img(sheet, rows, imgpath)
                isInsertImg = True
                continue
            sheet.cell(rows, data.index(d) + 1, d)
        wb.save(path)

    except PermissionError:
        print('保存数据失败！请关闭xlsx文件再运行！')


# 保存公司名称 生产模式
def saveData(path, row, data, url):
    try:
        wb = load_workbook(path)
        sheet = wb['Sheet']
        sheet.cell(row, 5, data[0])
        sheet.cell(row, 6, data[1])
        wb.save(path)
        donepath = cur_dir + r'\done\done.txt'
        f = open(donepath, 'a')
        f.write(url + '\n')
        f.close()
    except PermissionError:
        print('保存数据失败！请关闭xlsx文件再运行！')


# 加载产品图片，产品图片需要放在当前程序目录下的img文件夹
def loadImg():
    path = cur_dir + '\img' + '\\'
    donepath = cur_dir + r'\done\imgdone.txt'
    if os.path.exists(donepath):
        pass
    else:
        file = open(donepath, 'w')
        file.close()
    done = []
    f = open(donepath, 'r', encoding='utf8')
    line = f.readline()
    while line:
        done.append(line.strip())
        line = f.readline()
    f.close()
    for fileName in os.listdir(path):
        if fileName in done:
            continue
        url = 'https://www.1688.com'
        driver.get(url)
        inp = driver.find_element(By.CSS_SELECTOR, '.react-file-reader-input')
        inp.send_keys(path + fileName)
        time.sleep(3)
        imgName = fileName.rsplit(".", 1)[0]
        datapath = cur_dir + '\data\\' + imgName + '.xlsx'
        createXls(datapath)
        for i in range(PAGE):
            time.sleep(random.randint(1, 2))
            driver.execute_script("window.scrollBy(0,4000)")
        items = driver.find_elements(By.CSS_SELECTOR, '#sm-offer-list > div')
        donepath = cur_dir + r'\done\imgdone.txt'
        f = open(donepath, 'a', encoding='utf8')
        f.write(fileName + '\n')
        f.close()
        for item in items:
            try:
                href = item.find_element(By.CSS_SELECTOR, 'div > div.img-container > div > a')
                img = item.find_element(By.CSS_SELECTOR, 'div > div.img-container > div > a > div')
                price = item.find_element(By.CSS_SELECTOR, 'div > div.mojar-element-price > div.showPricec > div.price')
                sale = item.find_element(By.CSS_SELECTOR, 'div > div.mojar-element-price > div.sale > div')
                href = href.get_attribute('href').strip()
                price = price.text.strip()
                sale = sale.text.strip()
                img = img.get_attribute('style').strip()
                img = img.replace('background-image: url("', '')
                img = img.replace('");', '')
                saveTmpData(datapath, [img, price, sale, href])
            except:
                print('出错啦！请重新启动！')


# 商品详细
def detail():
    path = cur_dir + '\data' + '\\'
    donepath = cur_dir + r'\done\done.txt'
    if os.path.exists(donepath):
        pass
    else:
        file = open(donepath, 'w')
        file.close()
    done = []
    f = open(donepath, 'r')
    line = f.readline()
    while line:
        done.append(line.strip())
        line = f.readline()
    f.close()
    for fileName in os.listdir(path):
        filepath = path + fileName
        data = pd.read_excel(filepath)
        href = list(data['链接'])
        row = 2
        for url in href:
            try:
                if url in done:
                    row += 1
                    continue
                driver.get(url)
                time.sleep(2)
                div = WebDriverWait(driver, 60, 0.5).until(EC.presence_of_element_located(
                    (By.CSS_SELECTOR,
                     "#hd_0_container_0 > div:nth-child(1) > div:nth-child(2) > div > div:nth-child(1) > div:nth-child(3) > div > div:nth-child(1) > img")))
                ActionChains(driver).move_to_element(div).perform()
                time.sleep(1)
                i = 0
                while driver.page_source.find('经营模式') == -1:
                    if driver.page_source.find('所在地区') != -1 and driver.page_source.find(
                            '退换体验') != -1 and driver.page_source.find('品质体验') != -1:
                        donepath = cur_dir + r'\done\done.txt'
                        f = open(donepath, 'a')
                        f.write(url + '\n')
                        f.close()
                        break
                    if i == 5:
                        break
                    i += 1
                    driver.refresh()
                    div = WebDriverWait(driver, 30, 1).until(EC.presence_of_element_located(
                        (By.CSS_SELECTOR,
                         "#hd_0_container_0 > div:nth-child(1) > div:nth-child(2) > div > div:nth-child(1) > div:nth-child(3) > div > div:nth-child(1) > img")))
                    ActionChains(driver).move_to_element(div).perform()
                    time.sleep(1)
                if driver.page_source.find('经营模式') == -1:
                    row += 1
                    continue
                name = driver.find_element(By.CSS_SELECTOR,
                                           '#hd_0_container_0 > div:nth-child(1) > div:nth-child(2) > div > div:nth-child(1) > div:nth-child(3) > div > div:nth-child(1) > span').text.strip()
                manage = driver.find_element(By.CSS_SELECTOR,
                                             '#hd_0_container_0 > div:nth-child(1) > div:nth-child(2) > div > div:nth-child(1) > div:nth-child(4) > div')
                manageModel = manage.text.strip()
                manageModel = manageModel[manageModel.index('经营模式') + 4:manageModel.index('所在地区')]
                saveData(filepath, row, [name, manageModel], url)
                row += 1
            except:
                print('出错啦！请重新启动！')


def del_file(path):
    ls = os.listdir(path)
    for i in ls:
        c_path = os.path.join(path, i)
        if os.path.isdir(c_path):  # 如果是文件夹那么递归调用一下
            del_file(c_path)
        else:  # 如果是一个文件那么直接删除
            os.remove(c_path)


# 程序入口
if __name__ == '__main__':
    try:
        # cookiepath -- 如果频繁出现验证码可以去util文件夹尝试登录1688，我在测试过程中没有频繁出现验证码
        cookiepath = cur_dir.rsplit("\\", 1)[0] + '\\util\\1688.json'
        if os.path.exists(cookiepath):
            driver.get('https://www.1688.com')
            add_cookies(driver, cookiepath)
        createDir()
        loadImg()
        detail()
        # 将临时文件删了，避免占储存
        del_file(cur_dir + '\imgtmp')
        os.rmdir(cur_dir + '\imgtmp')
        driver.close()
    except:
        print('出错啦！请重新启动！')
